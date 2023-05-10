import pandas as pd

# ----------------------------------------------------------------------------------------------------------------------------------
# 'Item Name', 'Game Name' , 'Listed On' , 'Acted On', ' Display Price' ,  ' Price in Cents' , 
# ' Type' , ' Market Name' , ' App Id' , ' Context Id' , ' Asset Id' , ' Instance Id' , ' Class Id' , 
# ' Unowned Context Id' , ' Unowned Id' , ' Partner Name', ' Partner Link'

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, colors, fills, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows





# # ---------------------------------------------------------------------------------------------------------------
# column highlighting functions

def highlightPrice(data, worksheet):

    light_blue_fill = PatternFill(start_color="EEFEFF", end_color="EEFEFF", fill_type="solid")
    display_price_col_idx = data.columns.get_loc("Combined_Price") + 1

    for row in range(1, len(data) + 2):
        worksheet.cell(row=row, column=display_price_col_idx).fill = light_blue_fill
    

def highlightType(data, worksheet):

    red_fill = PatternFill(start_color="FFF4EE", end_color="FFF4EE", fill_type="solid")
    green_fill = PatternFill(start_color="EEFFF5", end_color="EEFFF5", fill_type="solid")

    col_idx = data.columns.get_loc(' Type') + 1
    for row in range(2, len(data) + 2):
        cell_value = worksheet.cell(row=row, column=col_idx).value
        if cell_value == 'purchase':
            for col in range(1, len(data.columns) + 1):
                worksheet.cell(row=row, column=col).fill = red_fill
        elif cell_value == 'sale':
            for col in range(1, len(data.columns) + 1):
                worksheet.cell(row=row, column=col).fill = green_fill
    

def highlightGame(data, worksheet):

    orange_fill = PatternFill(start_color="FFF5CC", end_color="FFF5CC", fill_type="solid")
    blue_fill = PatternFill(start_color="CCDAFF", end_color="CCDAFF", fill_type="solid")
    grey_fill = PatternFill(start_color="ADADAD", end_color="ADADAD", fill_type="solid")

    col_idx = data.columns.get_loc('Game Name') + 1
    for row in range(2, len(data) + 2):
        cell_value = worksheet.cell(row=row, column=col_idx).value
        if cell_value == 'TF2':
            worksheet.cell(row=row, column=col_idx).fill = orange_fill
        elif cell_value == 'CSGO':
            worksheet.cell(row=row, column=col_idx).fill = blue_fill
        elif 'Trading Card' in cell_value:
            worksheet.cell(row=row, column=col_idx).fill = grey_fill

# ---------------------------------------------------------------------------------------------------------------







# replace game names w/ abbreviations
def replaceGameNames(data):

    data['Game Name'] = data['Game Name'].replace('Counter-Strike: Global Offensive', 'CSGO', regex=True)
    data['Game Name'] = data['Game Name'].replace('Team Fortress 2', 'TF2', regex=True)

    return data



# combine bulk orders
def combineBulk(data):

    # clean up date columns
    data['Acted On'] = pd.to_datetime(data['Acted On']).dt.date
    data['Listed On'] = pd.to_datetime(data['Listed On']).dt.date

    # Create a new column to store the original index
    data['original_index'] = data.index

    # Group by 'Acted On' and 'Price in Cents', keep the first non-null value in each group, and calculate the count
    grouped_data = data.groupby(['Acted On', ' Price in Cents'], as_index=False).first()
    grouped_data['Count'] = data.groupby(['Acted On', ' Price in Cents']).size().values
    grouped_data['Combined_Price'] = grouped_data['Count'] * grouped_data[' Price in Cents'] / 100

    # Sort by the original index
    grouped_data = grouped_data.sort_values('original_index')

    # Drop the original index column
    grouped_data = grouped_data.drop(columns=['original_index'])

    # # clean up Acted On column
    # grouped_data['Acted On'] = grouped_data.to_datetime(grouped_data['Acted On']).dt.date

    # Drop unnecessary columns and reorder
    data = grouped_data[[
                                # 'Item Name', 
                                'Game Name', 
                                'Listed On', 
                                'Acted On', 
                                ' Display Price', 
                                # ' Price in Cents',
                                ' Type', 
                                ' Market Name', 
                                ' App Id', 
                                # ' Context Id', 
                                ' Asset Id', 
                                # ' Instance Id', 
                                ' Class Id',
                                # ' Unowned Context Id', 
                                ' Unowned Id', 
                                ' Partner Name', 
                                ' Partner Link', 
                                'Count', 
                                'Combined_Price']]
    
    
    return data


# fix input excel file naming format
def fixItemNames(data):
    data[' Market Name'] = data[' Market Name'].replace('â˜…', '★', regex=True)
    data[' Market Name'] = data[' Market Name'].replace('StatTrakâ„¢', 'StatTrak™', regex=True)

    return data





class marketItem:
    def __init__(self, index, name, game):
        self.index = index
        self.name = name
        self.game = game
















# https://docs.python.org/3/library/urllib.parse.html
import requests
def showCurrentPrice(data):

    itemName = "StatTrak™ XM1014 | Oxide Blaze (Minimal Wear)"

    gameID = "730"
    #    'CSGO' : 730
    #    'TF2' : 440
    #    To find any game's appid, view the store page in steam and check the URL

    currencyID = "20"
    #    'USD' : 1
    #    'GBP' : 2
    #    'EUR' : 3
    #    'CHF' : 4
    #    'RUB' : 5
    #    'KRW' : 16
    #    'CAD' : 20
    
    # GET /market/priceoverview
    urlConcat = "https://steamcommunity.com/market/priceoverview/?" + "currency=" + currencyID + "&appid=" + gameID + "&market_hash_name=" + itemName
    req = requests.get(url = urlConcat)

    
    # get all values of dict
    itemReq = req.json()

    if itemReq != None:
        itemIterator = iter(itemReq.values())
        itemQueryStatus, itemLowestPrice, itemVolume, itemMedianPice = next(itemIterator), next(itemIterator), next(itemIterator), next(itemIterator)
        
        # print values of dict
        print(itemReq)
        print(itemLowestPrice)
        print(itemVolume)
        print(itemMedianPice)


        data['Current_Price'] = itemLowestPrice
    else:
        print("itemReq = None -> API limit reached")

    return data









# need a way to bypass API limit - also, iterrows() takes extremely long
# https://docs.python.org/3/library/urllib.parse.html
import requests
import time
def automatedShowCurrentPrice(data):

    gameID = "730"

    currencyID = "20"


    # neat way of creating list of all entries in a column (not used in func)
    market_items = data.apply(lambda row: marketItem(row[' Class Id'], row[' Market Name'], row[' App Id']), axis=1).tolist()
    # print(market_items[1].name)


    for i, j in data.iterrows():

        # time.sleep(0)

        # GET /market/priceoverview
        print(i)
        urlConcat = "https://steamcommunity.com/market/priceoverview/?" + "currency=" + currencyID + "&appid=" + gameID + "&market_hash_name=" + data.loc[i, ' Market Name']
        req = requests.get(url = urlConcat)

        
        # get all values of dict
        itemReq = req.json()
        data.loc[i, 'currPrice'] = str(itemReq)

    

    return data




# ---------------------------------------------------------------------------------------------------------------
# writing output sheet

# OUTDATED !!!!!!!!
def createOutputSheet(data):
    # Save the resulting data to a new Excel file
    output_file_name = 'output_data_csgo_default.xlsx'
    writer = pd.ExcelWriter(output_file_name, engine='openpyxl')
    data.to_excel(writer, index=False, sheet_name='output1')
    writer.close()



def highlightSheet(output_file_name, data):

    # Save the DataFrame to Excel
    data.to_excel(output_file_name, index=False)

    # Load the workbook with openpyxl
    wb = load_workbook(output_file_name)
    ws = wb.active

    # Apply the highlights
    highlightType(data, ws)
    highlightGame(data, ws)

    # Save the modified Excel file
    wb.save(output_file_name)

# ---------------------------------------------------------------------------------------------------------------






def main():

    print("~~~ Welcome to CounterStrike Transac Analyzer ~~~\n")
    # inputSheetName = input("Please type the name of your input xlsx sheet (excl ext) >>")

    # print("\nHow would you like to convert your sheet?")
    # print("1 - smth1")
    # print("2 - smth2")
    # print("3 - smth3")

    # convertType = input(">>")

    file_name = 'steam_market_history.xlsx'
    sheet_name = 'steam_market_history'
    data = pd.read_excel(file_name, sheet_name=sheet_name)


    # data rewording
    data = replaceGameNames(data)
    data = fixItemNames(data)


    # combine bulk func call
    data = combineBulk(data)

    # show curr price func call
    # data = showCurrentPrice(data)

    

    # output normal doc
    # createOutputSheet(data)

    # output highlighted doc
    highlightSheet("output_data_csgo_final.xlsx", data)



    print("done")


if __name__ == "__main__":
    main()